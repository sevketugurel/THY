"""Unit tests for src.data.loaders — schema-validated readers for the 4 input files.

marker: unit (solver-free, pure IO/validation logic).
"""
import datetime as dt
from pathlib import Path

import openpyxl
import pytest

from src.data.loaders import (
    SchemaError,
    load_change_ranking,
    load_flight_pairs,
    load_od_table,
    load_yolcu_verisi,
)

FIXDIR = Path(__file__).parent.parent / "fixtures"
pytestmark = pytest.mark.unit


# ---------------------------------------------------------------------------
# O&D table
# ---------------------------------------------------------------------------
def test_load_od_table_returns_expected_row_count_and_columns():
    df = load_od_table(FIXDIR / "synthetic_od_table.xlsx")
    assert len(df) == 12 * 2 + 5  # 12 TK rows x 2 days + 5 rival rows
    assert list(df.columns) == [
        "cr1", "carrier_name", "dep1", "arr1", "flno1", "arr_time",
        "cr2", "dep2", "arr2", "flno2", "dep_time",
        "gate_to_gate_min", "od", "gun",
    ]


def test_load_od_table_normalizes_gate_to_gate_to_minutes():
    df = load_od_table(FIXDIR / "synthetic_od_table.xlsx")
    row = df[(df.flno1 == 9101) & (df.flno2 == 9112) & (df.gun == 1)].iloc[0]
    assert row.gate_to_gate_min == 280  # 220 (K_ZZA_ZZB) + 60 (gap)


def test_load_od_table_flno_columns_are_int():
    df = load_od_table(FIXDIR / "synthetic_od_table.xlsx")
    assert df.flno1.dtype.kind in "iu"
    assert df.flno2.dtype.kind in "iu"


def _write_od_table(tmp_path: Path, rows: list, header=None) -> Path:
    header = header or [
        "Cr1", "Carrier Name", "Dep1", "Arr1", "FlNo1", "Arr Time",
        "Cr2", "Dep2", "Arr2", "FlNo2", "Dep Time",
        "Gate-to-Gate Uçuş Süresi", "O&D", "Gün",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bağlantı Tablosu"
    ws.append(header)
    for r in rows:
        ws.append(r)
    path = tmp_path / "broken_od.xlsx"
    wb.save(path)
    return path


def test_load_od_table_rejects_cr1_cr2_mismatch(tmp_path):
    base = dt.datetime(2026, 1, 5, 10, 0)
    bad_row = [
        "TK", "Turkish Airlines", "ZZA", "IST", 1, base,
        "XX", "IST", "ZZB", 2, base + dt.timedelta(minutes=90),
        dt.time(2, 0), "ZZA-ZZB", 1,
    ]
    path = _write_od_table(tmp_path, [bad_row])
    with pytest.raises(SchemaError, match="Cr1"):
        load_od_table(path)


def test_load_od_table_rejects_dep2_arr1_mismatch(tmp_path):
    base = dt.datetime(2026, 1, 5, 10, 0)
    bad_row = [
        "TK", "Turkish Airlines", "ZZA", "IST", 1, base,
        "TK", "AMS", "ZZB", 2, base + dt.timedelta(minutes=90),
        dt.time(2, 0), "ZZA-ZZB", 1,
    ]
    path = _write_od_table(tmp_path, [bad_row])
    with pytest.raises(SchemaError, match="Dep2"):
        load_od_table(path)


def test_load_od_table_rejects_gun_out_of_range(tmp_path):
    base = dt.datetime(2026, 1, 5, 10, 0)
    bad_row = [
        "TK", "Turkish Airlines", "ZZA", "IST", 1, base,
        "TK", "IST", "ZZB", 2, base + dt.timedelta(minutes=90),
        dt.time(2, 0), "ZZA-ZZB", 8,
    ]
    path = _write_od_table(tmp_path, [bad_row])
    with pytest.raises(SchemaError, match="Gün"):
        load_od_table(path)


# ---------------------------------------------------------------------------
# Yolcu Verisi (rho)
# ---------------------------------------------------------------------------
def test_load_yolcu_verisi_returns_rho_lookup():
    df = load_yolcu_verisi(FIXDIR / "synthetic_yolcu_verisi.xlsx")
    assert len(df) == 2
    assert set(df.columns) == {"orig", "dest", "rho"}
    row = df[(df.orig == "ZZA") & (df.dest == "ZZB")].iloc[0]
    assert row.rho == 100


def test_load_yolcu_verisi_sums_duplicate_od_pairs(tmp_path):
    # Real data has ~12 duplicate (orig,dest) rows (confirmed against the full-data
    # file) -- treated as split/segmented rho contributions and summed, not rejected.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yolcu Verisi Son 1 Yıl"
    ws.append(["Orig Airport Code", "Dest Airport Code", "OD importance factor"])
    ws.append(["ZZA", "ZZB", 100])
    ws.append(["ZZA", "ZZB", 200])
    path = tmp_path / "dup_yv.xlsx"
    wb.save(path)
    df = load_yolcu_verisi(path)
    assert len(df) == 1
    assert df.iloc[0].rho == 300


def test_load_yolcu_verisi_rejects_missing_orig_or_dest(tmp_path):
    # Real full-data file has 3 rows with a null Dest Airport Code (confirmed by
    # inspection) -- these must fail loudly, not be silently dropped by groupby.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yolcu Verisi Son 1 Yıl"
    ws.append(["Orig Airport Code", "Dest Airport Code", "OD importance factor"])
    ws.append(["ZZA", None, 100])
    path = tmp_path / "missing_dest_yv.xlsx"
    wb.save(path)
    with pytest.raises(SchemaError, match="missing"):
        load_yolcu_verisi(path)


def test_load_yolcu_verisi_rejects_nonpositive_rho(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yolcu Verisi Son 1 Yıl"
    ws.append(["Orig Airport Code", "Dest Airport Code", "OD importance factor"])
    ws.append(["ZZA", "ZZB", 0])
    path = tmp_path / "broken_yv2.xlsx"
    wb.save(path)
    with pytest.raises(SchemaError, match="positive"):
        load_yolcu_verisi(path)


# ---------------------------------------------------------------------------
# change_ranking_input
# ---------------------------------------------------------------------------
def test_load_change_ranking_returns_lookup():
    df = load_change_ranking(FIXDIR / "synthetic_change_ranking_input.xlsx")
    assert set(df.columns) == {"n", "b", "r", "weight"}
    row = df[(df.n == 2) & (df.b == 2) & (df.r == 1)].iloc[0]
    assert row.weight == pytest.approx(1.6321205588285577)


def test_load_change_ranking_allows_negative_weight():
    df = load_change_ranking(FIXDIR / "synthetic_change_ranking_input.xlsx")
    row = df[(df.n == 3) & (df.b == 1) & (df.r == 2)].iloc[0]
    assert row.weight == pytest.approx(-0.132)


def test_load_change_ranking_rejects_rank_exceeding_n(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Rakip Sayısı", "İlk Rank", "Son Rank", "Weight"])
    ws.append([2, 1, 3, 0.5])  # r=3 > N=2, invalid
    path = tmp_path / "broken_rank.xlsx"
    wb.save(path)
    with pytest.raises(SchemaError, match="rank"):
        load_change_ranking(path)


# ---------------------------------------------------------------------------
# Flight Pairs
# ---------------------------------------------------------------------------
def test_load_flight_pairs_normalizes_padded_string_flno():
    df = load_flight_pairs(FIXDIR / "synthetic_flight_pairs.xlsx")
    assert df.flno.dtype.kind in "iu"
    assert 9311 in set(df.flno)


def test_load_flight_pairs_groups_by_pair_id():
    df = load_flight_pairs(FIXDIR / "synthetic_flight_pairs.xlsx")
    group = df[df.pair == "ROT-A"]
    assert set(group.flno) == {9311, 9301}
