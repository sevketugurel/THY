"""Solve tests for G (tarife düzenliliği / gün-içi tutarlılık) -- reference-time
formulation.

Doğruluk argümanı (ultrathink, kod öncesi): brief "aynı uçuş numarasının farklı
operasyon günlerindeki IST kalkış/varış saatleri birbirinden en fazla X_dev
dakika sapabilir" diyor -- yani max(t_h) - min(t_h) <= X_dev, TÜM günler h
üzerinden.

Standart (naif) formülasyon HER GÜN ÇİFTİ için |t_h1-t_h2|<=X_dev kurar:
O(H^2) kısıt. Kullanıcının istediği REFERANS-ZAMAN formülasyonu: serbest bir
T_f (flno'nun "referans" zamanı, yeni bir değişken), ve HER gün h için
T_f <= t_h <= T_f+X_dev. Bu O(H) kısıt (gün sayısı kadar üst+alt sınır).

Eşdeğerlik kanıtı: (<=) TÜM t_h aynı [T_f,T_f+X_dev] penceresindeyse, herhangi
iki günün farkı en fazla X_dev (ikisi de aynı X_dev-genişlik pencerede).
(=>) max(t_h)-min(t_h)<=X_dev ise, T_f=min(t_h) seçilir; o zaman TÜM
t_h<=max(t_h)<=min(t_h)+X_dev=T_f+X_dev otomatik sağlanır, ve t_h>=min(t_h)=T_f
zaten tanım gereği. Yani formülasyon TAM eşdeğer (gevşek değil), sadece daha
az kısıtla (O(H) vs O(H^2)) ifade ediyor -- daha küçük branch-and-bound ağacı.

Rol-ayrımı: aynı flno hem IB hem OB rolünde görünebilir (M1'de 26 örnek
doğrulandı) -- her ROL kendi ayrı G kısıtına tabi (brief "kalkış VE varış
saatleri" diyerek zaten ayrı ele alıyor).

marker: solve (small HiGHS solve, <60s).
"""
import datetime as dt
from pathlib import Path

import pandas as pd
import pyomo.environ as pyo
import pytest

from src.candidates.generate import Candidate, compute_epoch_anchor, generate_candidates
from src.data.loaders import load_od_table
from src.model.constraints_operations import add_g_constraints
from src.model.constraints_selection import add_b_constraints, add_c_constraints, add_flight_time_variables
from src.solve.runner import solve

FIXDIR = Path(__file__).parent.parent / "fixtures"
pytestmark = pytest.mark.solve

L, U = 60, 300
X_DEV = 15


@pytest.fixture
def tk_free():
    od_table = load_od_table(FIXDIR / "synthetic_od_table.xlsx")
    tk = od_table[od_table.cr1 == "TK"]
    anchor = compute_epoch_anchor(tk)
    candidates = []
    for gun in sorted(int(g) for g in tk["gun"].unique()):
        candidates.extend(generate_candidates(
            tk, L=L, U=U, gun=gun, adjustable_window_min=180, adjustable_set="all",
            epoch_anchor=anchor,
        ))
    return candidates, anchor


def _build(candidates, anchor, x_dev):
    model = pyo.ConcreteModel()
    add_flight_time_variables(model, candidates)
    add_b_constraints(model, candidates, L=L, U=U)
    add_c_constraints(model, candidates)
    add_g_constraints(model, candidates, anchor, x_dev)
    model._candidates = candidates
    return model


def test_g_binding_forces_mi1_within_x_dev_across_days(tk_free):
    # MI1(9101, IB role) baseline: Gün1 arr=840, Gün2 arr=815 -- SAME TIME OF
    # DAY diff=25>15, violates at baseline. With free times (+-180 window),
    # the solver MUST shift at least one day to bring them within 15 of each
    # other. Gün2 is exactly 1 calendar day after Gün1 in this fixture (see
    # tests/fixtures/build_fixture.py GUN2_BASE=GUN1_BASE+1 day), so raw
    # epoch-minute values differ by ~1440 even when compliant -- subtract that
    # day offset before comparing (matches add_g_constraints's own
    # day-normalization, see constraints_operations.py::_day_offsets).
    candidates, anchor = tk_free
    model = _build(candidates, anchor, x_dev=X_DEV)
    model.objective = pyo.Objective(expr=0, sense=pyo.maximize)  # feasibility-only
    result = solve(model, solver="highs", time_limit_sec=60, seed=42)
    assert result.status == "optimal"
    t1 = pyo.value(model.t_arr["IB", 9101, 1])
    t2 = pyo.value(model.t_arr["IB", 9101, 2])
    assert abs(t2 - 1440 - t1) <= X_DEV


def test_g_non_binding_leaves_ni1_unrestricted(tk_free):
    # NI1(9201, IB role) baseline: Gün1 arr=795, Gün2 arr=800 -- diff=5<=15,
    # already compliant. Objective pushes t_arr[Gün1] to its window max;
    # G must not additionally restrict it below what B/window already allow.
    candidates, anchor = tk_free
    model = _build(candidates, anchor, x_dev=X_DEV)
    model.objective = pyo.Objective(expr=model.t_arr["IB", 9201, 1], sense=pyo.maximize)
    result = solve(model, solver="highs", time_limit_sec=60, seed=42)
    assert result.status == "optimal"
    # window is baseline(795)+-180 -> max 975. Gün2 has its own +-180 window
    # (800+-180=[620,980]) with enough day-adjusted room to accommodate
    # Gün1=975 without G becoming an additional bottleneck.
    assert pyo.value(model.t_arr["IB", 9201, 1]) == pytest.approx(975.0)


def test_g_violation_caught_by_validator():
    import json
    import tempfile

    from src.validate.independent_validator import validate_output

    data = {
        "objective_value": 0.0,
        "selected_connections": [],
        "adjusted_flight_times": [
            # Global epoch values (matching real output format): Gün1=840
            # (14:00 on day 0), Gün2=1440+815=2255 (13:35 on day 1) -- day-of-day
            # diff=25>15, genuinely violates.
            {"role": "IB", "flno": 9101, "gun": 1, "time_min": 840},
            {"role": "IB", "flno": 9101, "gun": 2, "time_min": 1440 + 815},
        ],
        "ranking_results": [],
        "solver_metrics": {"status": "optimal", "solve_time_sec": 0.1},
    }
    with tempfile.TemporaryDirectory() as d:
        path = Path(d) / "output.json"
        path.write_text(json.dumps(data))
        result = validate_output(
            path, FIXDIR / "synthetic_od_table.xlsx", L=L, U=U,
            adjustable_window_min=180, adjustable_set="all", x_dev=X_DEV,
        )
        assert not result.is_valid
        assert any("x_dev" in v.lower() or "düzenlilik" in v.lower() or "regularity" in v.lower()
                   for v in result.violations)


# --- M4 "G check": gece yarısı sarma (midnight wraparound) ---

def _midnight_candidate(gun, arr_tod_min, anchor, flno=9301):
    # Rfix (arr_lo=arr_hi=arr_tod_min-anchored epoch) -- deterministic,
    # isolates the test to G's own day-offset arithmetic (no solver freedom
    # to dodge a broken formulation).
    ts = anchor + pd.Timedelta(days=gun - 1, minutes=arr_tod_min)
    epoch = int((ts - anchor).total_seconds() // 60)
    return Candidate(
        od="ZZM-IST", o="ZZM", d="IST", gun=gun, flno1=flno, flno2=80000 + gun,
        r1_id=("IB", flno, gun), r2_id=("OB", 80000 + gun, gun),
        arr_time=ts, dep_time=ts, gap_min=0,
        arr_lo=epoch, arr_hi=epoch, dep_lo=0, dep_hi=0,
        gap_lo=-epoch, gap_hi=-epoch,
    )


def test_g_no_false_violation_at_midnight_wraparound():
    # Gün1 baseline arr=23:55 (day-of-day=1435), Gün2 baseline arr=00:05
    # (day-of-day=5) -- REAL clock difference is 10min, well within
    # X_dev=15. Pure-midnight day-offset anchoring would normalize these to
    # 1435 and 5 -- a FAKE ~1430min spread that makes G (wrongly)
    # infeasible even though the true schedule is fully compliant.
    anchor = pd.Timestamp("2024-01-01")
    c1 = _midnight_candidate(gun=1, arr_tod_min=1435, anchor=anchor)
    c2 = _midnight_candidate(gun=2, arr_tod_min=5, anchor=anchor)
    model = pyo.ConcreteModel()
    add_flight_time_variables(model, [c1, c2])
    add_g_constraints(model, [c1, c2], anchor, x_dev=15)
    model._candidates = [c1, c2]
    model.objective = pyo.Objective(expr=0, sense=pyo.maximize)
    result = solve(model, solver="highs", time_limit_sec=60, seed=42)
    assert result.status == "optimal"


def test_g_genuinely_far_apart_occurrences_split_into_separate_clusters():
    # Gün1 baseline arr=23:00 (1380), Gün2 baseline arr=01:00 next day
    # (1440+60=1500) -- REAL clock difference is 120min > X_dev=15, a
    # genuine (non-wraparound) mismatch. Under M5's clustered-G (VARSAYIM-9),
    # this does NOT make the model infeasible anymore -- it simply forms TWO
    # separate reconcilable clusters (Rfix, window=0 each, so the
    # reconcilability limit is 0+0+15=15 < 120), and G imposes NO
    # cross-cluster regularity. The wraparound fix must not swallow real
    # near-midnight mismatches into a false single cluster either -- this is
    # a genuine 120min gap, not a ~10min circular one.
    anchor = pd.Timestamp("2024-01-01")
    c1 = _midnight_candidate(gun=1, arr_tod_min=1380, anchor=anchor, flno=9302)
    c2 = _midnight_candidate(gun=2, arr_tod_min=60, anchor=anchor, flno=9302)
    model = pyo.ConcreteModel()
    add_flight_time_variables(model, [c1, c2])
    add_g_constraints(model, [c1, c2], anchor, x_dev=15)
    model._candidates = [c1, c2]
    model.objective = pyo.Objective(expr=0, sense=pyo.maximize)
    result = solve(model, solver="highs", time_limit_sec=60, seed=42)
    # A fully-fixed (Rfix), unconstrained-by-anything, trivial-objective model
    # can report "unknown" rather than "optimal" from HiGHS -- the important
    # thing is it's NOT infeasible (no false G violation) and no constraint
    # was built at all (both occurrences became singleton clusters).
    assert result.status != "infeasible"
    assert len(model.G_FLIGHTS) == 0


def test_g_within_cluster_regularity_still_binds_despite_distant_outlier():
    # VARSAYIM-9's exact TK2841 shape: 3 occurrences at ~03:25 (close, must
    # stay regular) + 1 at 14:10 (a genuine outlier, its own cluster).
    # Adjustable window +-30min (small, so the adversarial objective can't
    # trivially satisfy X_dev by accident) -- confirms the CLOSE trio is
    # still forced within a common X_dev=15 band even though a structurally
    # incompatible 4th occurrence exists for the SAME flight number.
    anchor = pd.Timestamp("2024-01-01")
    w = 30

    def _adjustable_candidate(gun, arr_tod_min, flno=9303):
        ts = anchor + pd.Timedelta(days=gun - 1, minutes=arr_tod_min)
        epoch = int((ts - anchor).total_seconds() // 60)
        return Candidate(
            od="ZZM-IST", o="ZZM", d="IST", gun=gun, flno1=flno, flno2=80000 + gun,
            r1_id=("IB", flno, gun), r2_id=("OB", 80000 + gun, gun),
            arr_time=ts, dep_time=ts, gap_min=0,
            arr_lo=epoch - w, arr_hi=epoch + w, dep_lo=0, dep_hi=0,
            gap_lo=-(epoch + w), gap_hi=-(epoch - w),
        )

    c1 = _adjustable_candidate(gun=1, arr_tod_min=205)   # 03:25
    c2 = _adjustable_candidate(gun=2, arr_tod_min=205)
    c3 = _adjustable_candidate(gun=3, arr_tod_min=205)
    # Rfix outlier (window=0, matching _midnight_candidate's pattern) --
    # avoids B's Big-M discipline entirely (irrelevant to this test) while
    # still exercising "does a genuinely-unreconcilable 4th occurrence get
    # its own singleton cluster instead of breaking the close trio's G".
    c_outlier = _midnight_candidate(gun=5, arr_tod_min=850, anchor=anchor, flno=9303)  # 14:10

    model = pyo.ConcreteModel()
    all_c = [c1, c2, c3, c_outlier]
    add_flight_time_variables(model, all_c)
    add_g_constraints(model, all_c, anchor, x_dev=15)
    model._candidates = all_c
    # Adversarial: push gün1 and gün3's arr as far apart as their own windows
    # allow (would violate X_dev=15 within the {1,2,3} cluster if G didn't bind).
    model.objective = pyo.Objective(
        expr=model.t_arr["IB", 9303, 3] - model.t_arr["IB", 9303, 1], sense=pyo.maximize,
    )
    result = solve(model, solver="highs", time_limit_sec=60, seed=42)
    assert result.status == "optimal"
    # gün3's raw t_arr is ~2 calendar days (2*1440min) ahead of gün1's --
    # maximizing the RAW difference is strategically equivalent to
    # maximizing the day-normalized one (a fixed constant offset doesn't
    # change which solution is optimal), but the ASSERTION must day-normalize
    # to compare like with like (same lesson as G's own _day_offsets).
    raw_diff = pyo.value(model.t_arr["IB", 9303, 3]) - pyo.value(model.t_arr["IB", 9303, 1])
    day_normalized_diff = raw_diff - 2 * 1440  # gün3 is 2 calendar days after gün1
    assert day_normalized_diff == pytest.approx(15.0), \
        "G must still cap the {1,2,3} cluster's spread at X_dev=15"
    assert len(model.G_FLIGHTS) == 1  # only the {1,2,3} cluster needs a constraint; the outlier is its own singleton


def test_validate_no_false_x_dev_violation_at_midnight_wraparound(tmp_path):
    # Same scenario as the model-side test, verified independently through
    # the validator: TK flno=9301 baseline arr=23:55 (Gün1) / 00:05 (Gün2)
    # -- real clock difference 10min, well within X_dev=15. A self-contained
    # tiny od_table fixture (the shared synthetic fixture has no near-midnight
    # flights) with both legs present so load_od_table's own invariants
    # (Cr1==Cr2, Dep2==Arr1) are satisfied.
    import json

    import openpyxl

    from src.data.loaders import load_od_table
    from src.validate.independent_validator import validate_output

    # pandas.DataFrame.to_excel() round-trips datetime.time/Timedelta cells
    # as plain strings/floats through openpyxl in this environment (verified
    # -- not a real-data concern, since load_od_table works correctly against
    # the checked-in fixtures/real files, which are written via a direct
    # openpyxl.Workbook() like this one, matching tests/fixtures/build_fixture.py's
    # own convention).
    anchor = dt.datetime(2024, 1, 1)
    arr1 = anchor + dt.timedelta(minutes=1435)  # Gün1 23:55
    arr2 = anchor + dt.timedelta(days=1, minutes=5)  # Gün2 00:05

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Cr1", "Carrier Name", "Dep1", "Arr1", "FlNo1", "Arr Time",
               "Cr2", "Dep2", "Arr2", "FlNo2", "Dep Time",
               "Gate-to-Gate Uçuş Süresi", "O&D", "Gün"])
    ws.append(["TK", "Turkish Airlines", "ZZM", "IST", 9301, arr1,
               "TK", "IST", "ZZN", 80001, arr1 + dt.timedelta(minutes=100),
               dt.time(3, 20), "ZZM-ZZN", 1])
    ws.append(["TK", "Turkish Airlines", "ZZM", "IST", 9301, arr2,
               "TK", "IST", "ZZN", 80002, arr2 + dt.timedelta(minutes=100),
               dt.time(3, 20), "ZZM-ZZN", 2])
    od_path = tmp_path / "od_table.xlsx"
    wb.save(od_path)
    load_od_table(od_path)  # sanity: fixture satisfies the loader's own invariants

    output_path = tmp_path / "output.json"
    output_path.write_text(json.dumps({
        "objective_value": 0.0,
        "selected_connections": [],
        "adjusted_flight_times": [
            {"role": "IB", "flno": 9301, "gun": 1, "time_min": 1435},
            {"role": "IB", "flno": 9301, "gun": 2, "time_min": 1445},
        ],
        "ranking_results": [],
        "solver_metrics": {"status": "optimal", "solve_time_sec": 0.1},
    }))
    result = validate_output(
        output_path, od_path, L=L, U=U,
        adjustable_window_min=180, adjustable_set="all", x_dev=15,
    )
    assert not any("x_dev" in v.lower() or "regularity" in v.lower() for v in result.violations), result.violations


def _clustering_fixture_od_table(tmp_path):
    # TK flno=9401: gün1/2/3 baseline ~03:25 (close, reconcilable at
    # baseline), gün5 baseline 14:10 (genuine outlier, own cluster).
    # adjustable_window_min=30 -> close trio's reconcilability limit is
    # 30+30+15=75, comfortably covering their ~10min baseline spread; the
    # outlier's baseline gap (~645min) is far beyond any window.
    import openpyxl

    anchor = dt.datetime(2024, 1, 1)
    rows = [
        (9401, 1, anchor + dt.timedelta(minutes=205), 80101),   # 03:25
        (9401, 2, anchor + dt.timedelta(days=1, minutes=210), 80102),  # 03:30
        (9401, 3, anchor + dt.timedelta(days=2, minutes=200), 80103),  # 03:20
        (9401, 5, anchor + dt.timedelta(days=4, minutes=850), 80105),  # 14:10, outlier
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Cr1", "Carrier Name", "Dep1", "Arr1", "FlNo1", "Arr Time",
               "Cr2", "Dep2", "Arr2", "FlNo2", "Dep Time",
               "Gate-to-Gate Uçuş Süresi", "O&D", "Gün"])
    for flno1, gun, arr, flno2 in rows:
        ws.append(["TK", "Turkish Airlines", "ZZM", "IST", flno1, arr,
                   "TK", "IST", "ZZN", flno2, arr + dt.timedelta(minutes=100),
                   dt.time(3, 20), "ZZM-ZZN", gun])
    od_path = tmp_path / "od_table_clustering.xlsx"
    wb.save(od_path)
    load_od_table(od_path)
    return od_path


def test_validate_catches_within_cluster_violation_alongside_unreconcilable_outlier(tmp_path):
    import json

    from src.validate.independent_validator import validate_output

    od_path = _clustering_fixture_od_table(tmp_path)
    # Reported times: gün1 shifted to 225, gün2 shifted to 190, gün3 at 200
    # (all within their own +-30min window) -- spread=225-190=35>X_dev=15,
    # a GENUINE within-cluster violation (the trio IS reconcilable at
    # baseline, so G's cluster requires them to actually comply). Outlier
    # (gün5) reported at its own baseline (850) -- must NOT be flagged
    # despite being wildly different from the trio's times.
    output_path = tmp_path / "output.json"
    output_path.write_text(json.dumps({
        "objective_value": 0.0,
        "selected_connections": [],
        "adjusted_flight_times": [
            {"role": "IB", "flno": 9401, "gun": 1, "time_min": 225},
            {"role": "IB", "flno": 9401, "gun": 2, "time_min": 1440 + 190},
            {"role": "IB", "flno": 9401, "gun": 3, "time_min": 2880 + 200},
            {"role": "IB", "flno": 9401, "gun": 5, "time_min": 4 * 1440 + 850},
        ],
        "ranking_results": [],
        "solver_metrics": {"status": "optimal", "solve_time_sec": 0.1},
    }))
    result = validate_output(
        output_path, od_path, L=L, U=U,
        adjustable_window_min=30, adjustable_set="all", x_dev=15,
    )
    x_dev_violations = [v for v in result.violations if "x_dev" in v.lower() or "regularity" in v.lower()]
    assert len(x_dev_violations) == 1, x_dev_violations
    assert "9401" in x_dev_violations[0]
    # The outlier (gün5) must not appear in the flagged cluster.
    assert "5" not in x_dev_violations[0].split("küme=")[1].split(":")[0]


def test_validate_allows_reconcilable_trio_when_within_x_dev_despite_outlier(tmp_path):
    import json

    from src.validate.independent_validator import validate_output

    od_path = _clustering_fixture_od_table(tmp_path)
    # Same shape, but the trio's reported spread is now within X_dev=15 --
    # no violation, regardless of the outlier's reported time.
    output_path = tmp_path / "output.json"
    output_path.write_text(json.dumps({
        "objective_value": 0.0,
        "selected_connections": [],
        "adjusted_flight_times": [
            {"role": "IB", "flno": 9401, "gun": 1, "time_min": 205},
            {"role": "IB", "flno": 9401, "gun": 2, "time_min": 1440 + 210},
            {"role": "IB", "flno": 9401, "gun": 3, "time_min": 2880 + 200},
            {"role": "IB", "flno": 9401, "gun": 5, "time_min": 4 * 1440 + 850},
        ],
        "ranking_results": [],
        "solver_metrics": {"status": "optimal", "solve_time_sec": 0.1},
    }))
    result = validate_output(
        output_path, od_path, L=L, U=U,
        adjustable_window_min=30, adjustable_set="all", x_dev=15,
    )
    assert not any("x_dev" in v.lower() or "regularity" in v.lower() for v in result.violations), result.violations
