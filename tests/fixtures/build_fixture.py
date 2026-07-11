"""
Builds the synthetic test fixtures (4 xlsx files mirroring the real schemas exactly).

This is a ONE-TIME deterministic build script, not a data generator: every value
written here was hand-chosen and hand-verified (see README.md for the full derivation
and every hand-calculated expected result). Re-running this script always produces
byte-for-byte identical output. Run with: python tests/fixtures/build_fixture.py
"""
import datetime as dt
from pathlib import Path

import openpyxl

FIXDIR = Path(__file__).parent

GUN1_BASE = dt.date(2026, 1, 5)
GUN2_BASE = dt.date(2026, 1, 6)


def dtm(base_date: dt.date, minutes: int) -> dt.datetime:
    return dt.datetime.combine(base_date, dt.time(0, 0)) + dt.timedelta(minutes=minutes)


def duration_min(m: int) -> dt.time:
    h, mm = divmod(m, 60)
    return dt.time(h, mm)


# ---------------------------------------------------------------------------
# O&D Rakip Bağlantı Tablosu.xlsx
# ---------------------------------------------------------------------------
HEADER_OD = [
    "Cr1", "Carrier Name", "Dep1", "Arr1", "FlNo1", "Arr Time",
    "Cr2", "Dep2", "Arr2", "FlNo2", "Dep Time",
    "Gate-to-Gate Uçuş Süresi", "O&D", "Gün",
]

K_ZZA_ZZB = 220  # T_IB_ZZA(120) + T_OB_ZZB(100)
K_ZZB_ZZA = 240  # T_IB_ZZB(110) + T_OB_ZZA(130)
PLACEHOLDER_GTG = 200  # unused by block_times (gap invalid -> row excluded from K_od median)


def od_row(flno1, arr_min, flno2, dep_min, base_date, o, d, gtg_min):
    return [
        "TK", "Turkish Airlines", o, "IST", flno1, dtm(base_date, arr_min),
        "TK", "IST", d, flno2, dtm(base_date, dep_min),
        duration_min(gtg_min), f"{o}-{d}", 1 if base_date == GUN1_BASE else 2,
    ]


def rival_row(od, t_comp_min, flno1, flno2, carrier, gun=1):
    # carrier = distinct Cr1 code -- a "rival" is a CARRIER (brief's N_od /
    # T_comp_k semantics), not a row/itinerary. Multiple rows under the SAME
    # carrier consolidate to that carrier's best (min) itinerary.
    base = GUN1_BASE if gun == 1 else GUN2_BASE
    o, d = od.split("-")
    return [
        carrier, f"Rival {carrier}", o, "XHB", flno1, dtm(base, 600),
        carrier, "XHB", d, flno2, dtm(base, 600 + 90),
        duration_min(t_comp_min), od, gun,
    ]


def build_od_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bağlantı Tablosu"
    ws.append(HEADER_OD)

    for base_date, mi1_arr, ni1_arr in [(GUN1_BASE, 840, 795), (GUN2_BASE, 815, 800)]:
        gap_mi1_mo2 = 900 - mi1_arr
        gap_ni1_no2 = 1000 - ni1_arr
        rows = [
            od_row(9101, mi1_arr, 9111, 480, base_date, "ZZA", "ZZB", PLACEHOLDER_GTG),           # MI1xMO1 invalid
            od_row(9101, mi1_arr, 9112, 900, base_date, "ZZA", "ZZB", K_ZZA_ZZB + gap_mi1_mo2),    # MI1xMO2 valid
            od_row(9102, 600, 9111, 480, base_date, "ZZA", "ZZB", PLACEHOLDER_GTG),                 # MI2xMO1 invalid
            od_row(9102, 600, 9112, 900, base_date, "ZZA", "ZZB", K_ZZA_ZZB + 300),                 # MI2xMO2 valid
            od_row(9201, ni1_arr, 9211, 480, base_date, "ZZB", "ZZA", PLACEHOLDER_GTG),             # NI1xNO1 invalid
            od_row(9201, ni1_arr, 9212, 1000, base_date, "ZZB", "ZZA", K_ZZB_ZZA + gap_ni1_no2),    # NI1xNO2 valid
            od_row(9202, 500, 9211, 480, base_date, "ZZB", "ZZA", PLACEHOLDER_GTG),                 # NI2xNO1 invalid
            od_row(9202, 500, 9212, 1000, base_date, "ZZB", "ZZA", PLACEHOLDER_GTG),                # NI2xNO2 invalid (>U)
            od_row(9202, 500, 9311, 200, base_date, "ZZB", "ZZA", PLACEHOLDER_GTG),                 # discoverability: RA1
            od_row(9301, 850, 9111, 480, base_date, "ZZA", "ZZB", PLACEHOLDER_GTG),                 # discoverability: RA2
            od_row(9102, 600, 9411, 300, base_date, "ZZA", "ZZB", PLACEHOLDER_GTG),                 # discoverability: RB1
            od_row(9401, 555, 9211, 480, base_date, "ZZB", "ZZA", PLACEHOLDER_GTG),                 # discoverability: RB2
        ]
        for r in rows:
            ws.append(r)

    # rivals (Gün=1 only; ranking/D is independent of the G day-to-day test).
    # Each rival is a DISTINCT carrier (N_od = distinct Cr1 count, per brief's
    # semantics) -- ZZA-ZZB has N=2 (R1,R2), ZZB-ZZA has N=3 (R3,R4,R5).
    rivals = [
        rival_row("ZZA-ZZB", 300, 8001, 8101, carrier="R1"),  # beaten by MI1xMO2 (J=280)
        rival_row("ZZA-ZZB", 250, 8002, 8102, carrier="R2"),  # not beaten
        rival_row("ZZB-ZZA", 500, 8003, 8103, carrier="R3"),  # beaten by NI1xNO2 (J=445)
        rival_row("ZZB-ZZA", 400, 8004, 8104, carrier="R4"),  # not beaten (best itinerary)
        rival_row("ZZB-ZZA", 450, 8014, 8114, carrier="R4"),  # R4's WORSE 2nd itinerary -- must
                                                                # consolidate to min(400,450)=400,
                                                                # NOT count as a 4th rival
        rival_row("ZZB-ZZA", 445, 8005, 8105, carrier="R5"),  # beaten (boundary, J==T_comp)
    ]
    for r in rivals:
        ws.append(r)

    wb.save(FIXDIR / "synthetic_od_table.xlsx")


# ---------------------------------------------------------------------------
# Yolcu Verisi_masked.xlsx
# ---------------------------------------------------------------------------
def build_yolcu_verisi():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yolcu Verisi Son 1 Yıl"
    ws.append(["Orig Airport Code", "Dest Airport Code", "OD importance factor"])
    ws.append(["ZZA", "ZZB", 100])
    ws.append(["ZZB", "ZZA", 50])
    wb.save(FIXDIR / "synthetic_yolcu_verisi.xlsx")


# ---------------------------------------------------------------------------
# change_ranking_input.xlsx
# ---------------------------------------------------------------------------
def build_ranking_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Rakip Sayısı", "İlk Rank", "Son Rank", "Weight"])
    rows = [
        (1, 1, 1, 0.0),
        (2, 1, 1, 1.0),
        (2, 1, 2, 0.36787944117144233),
        (2, 2, 1, 1.6321205588285577),   # real value, reused verbatim from actual data
        (2, 2, 2, 1.0),
        (3, 1, 1, 0.0),
        (3, 1, 2, -0.132),                # matches brief's own illustrative example
        (3, 1, 3, -0.5),
    ]
    for r in rows:
        ws.append(list(r))
    wb.save(FIXDIR / "synthetic_change_ranking_input.xlsx")


# ---------------------------------------------------------------------------
# Flight Pairs.xlsx
# ---------------------------------------------------------------------------
def build_flight_pairs():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "W26 Flight Pair Synthetic"
    ws.append(["FlNo", "Orig", "Dest", "Pair"])
    rows = [
        (" 9311", "IST", "ZZA", "ROT-A"),  # RA1 outbound (non-binding rotation)
        (" 9301", "ZZA", "IST", "ROT-A"),  # RA2 inbound
        (" 9411", "IST", "ZZB", "ROT-B"),  # RB1 outbound (binding rotation, exact equality)
        (" 9401", "ZZB", "IST", "ROT-B"),  # RB2 inbound
    ]
    for r in rows:
        ws.append(list(r))
    wb.save(FIXDIR / "synthetic_flight_pairs.xlsx")


# ---------------------------------------------------------------------------
# synthetic_od_table_elapsed.xlsx (M5e, VARSAYIM-14/15) -- v2 schema, exercises
# the ElapsedTime1/ML2/ElapsedTime2 string-parse primary path + wrap-fix +
# [L,U]-independent K_od median, at real-loader (xlsx round-trip) scale.
# Does NOT touch synthetic_od_table.xlsx -- that fixture stays byte-identical
# and continues to exercise the LS-fallback (no Elapsed columns) path.
# ---------------------------------------------------------------------------
HEADER_OD_ELAPSED = [
    "Cr1", "Carrier Name", "Dep1", "Arr1", "FlNo1", "ElapsedTime1", "Arr Time",
    "Cr2", "Dep2", "Arr2", "FlNo2", "ML2", "ElapsedTime2", "Dep Time",
    "Gate-to-Gate Uçuş Süresi", "O&D", "Gün",
]


def _elapsed_str(minutes: int) -> str:
    # Mirrors the real v2 file's Excel-1899-epoch string artifact exactly
    # (docs/decisions.md 2026-07-11 M5e entry) -- primary parse path.
    h, m = divmod(minutes, 60)
    return f"30.12.1899 {h:02d}:{m:02d}:00"


def od_row_elapsed(flno1, elapsed1_min, arr_min, flno2, elapsed2_min, dep_min, base_date, o, d, ml2="M"):
    gap_min = dep_min - arr_min
    true_total = elapsed1_min + gap_min + elapsed2_min
    displayed = true_total % 1440  # organizer's wrap bug, reproduced deliberately
    return [
        "TK", "Turkish Airlines", o, "IST", flno1, _elapsed_str(elapsed1_min), dtm(base_date, arr_min),
        "TK", "IST", d, flno2, ml2, _elapsed_str(elapsed2_min), dtm(base_date, dep_min),
        duration_min(displayed), f"{o}-{d}", 1 if base_date == GUN1_BASE else 2,
    ]


# Market ZZA-ZZB: 3 TK rows, one with an out-of-[L,U] gap (L=60,U=300) that
# VARSAYIM-15 says must still count toward the K_od median (unlike the
# legacy LS path, which would exclude it -- see test_block_times.py's
# elapsed-path tests for the equivalent raw-DataFrame proof).
#   Row A: elapsed1=90,  gap=100 (valid),        elapsed2=130 -> k=220
#   Row B: elapsed1=100, gap=150 (valid),        elapsed2=130 -> k=230
#   Row C: elapsed1=50,  gap=40  (INVALID, <L),  elapsed2=150 -> k=200
#   median(220, 230, 200) = 220 (NOT 225, which is median(220,230) if Row C
#   were wrongly excluded).
K_OD_ELAPSED_ZZA_ZZB = 220

# Market EZE-PEK: real hand-verified wrap oracle (docs/decisions.md
# 2026-07-11 M5e entry) -- elapsed1=1020, gap=155 (valid), elapsed2=545 ->
# true total 1720min (28h40m); displayed (wrapped) field is 280min
# (datetime.time(4,40)) and must NOT be what gets used.
TRUE_TOTAL_EZE_PEK = 1720


def build_od_table_elapsed():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bağlantı Tablosu"
    ws.append(HEADER_OD_ELAPSED)

    rows = [
        od_row_elapsed(9501, 90, 700, 9502, 130, 800, GUN1_BASE, "ZZA", "ZZB"),   # Row A, k=220
        od_row_elapsed(9503, 100, 700, 9504, 130, 850, GUN1_BASE, "ZZA", "ZZB"),  # Row B, k=230
        od_row_elapsed(9505, 50, 700, 9506, 150, 740, GUN1_BASE, "ZZA", "ZZB"),   # Row C, k=200, gap=40<L
        od_row_elapsed(9601, 1020, 600, 9602, 545, 755, GUN1_BASE, "EZE", "PEK"), # wrap oracle, gap=155
    ]
    for r in rows:
        ws.append(r)

    wb.save(FIXDIR / "synthetic_od_table_elapsed.xlsx")


if __name__ == "__main__":
    build_od_table()
    build_yolcu_verisi()
    build_ranking_table()
    build_flight_pairs()
    build_od_table_elapsed()
    print("Fixtures written to", FIXDIR)
